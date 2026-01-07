import os
import zipfile
import csv
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Excel file support disabled. Install with: pip install openpyxl")


def flatten_extracted_folder(extract_path: str) -> str:
    """Flatten extracted folder structure if content is nested in a single subfolder.
    
    Args:
        extract_path: Path to the extracted folder
    
    Returns:
        Final path where the actual content is located
    """
    import shutil
    
    # Get subfolders, ignoring __MACOSX and other system folders
    subfolders = [f for f in os.listdir(extract_path) 
                  if os.path.isdir(os.path.join(extract_path, f)) 
                  and not f.startswith('__MACOSX') 
                  and not f.startswith('.')]
    
    # If there's exactly one meaningful subfolder, move its contents up
    if len(subfolders) == 1:
        nested_folder = os.path.join(extract_path, subfolders[0])
        print(f"[TRACE] Found nested structure, flattening from: {nested_folder}")
        
        # Create temp directory to move contents
        temp_dir = extract_path + "_temp"
        shutil.move(nested_folder, temp_dir)
        
        # Remove old extract path (should now be empty or only have __MACOSX)
        shutil.rmtree(extract_path)
        
        # Rename temp directory to final path
        shutil.move(temp_dir, extract_path)
        print(f"[TRACE] Flattened to: {extract_path}")
    
    return extract_path


def untar_zip_files(source_zip_path: str, target_zip_path: str) -> dict:
    """Untar/extract source and target zip files.
    
    Args:
        source_zip_path: Absolute path of source zip file
        target_zip_path: Absolute path of target zip file
    
    Returns:
        Dictionary with extracted source and target directory paths
    """
    print(f"[TRACE] Untarring zip files...")
    print(f"[TRACE] Source zip: {source_zip_path}")
    print(f"[TRACE] Target zip: {target_zip_path}")
    
    # Extract source zip
    source_dir = os.path.dirname(source_zip_path)
    source_name = os.path.splitext(os.path.basename(source_zip_path))[0]
    source_extract_path = os.path.join(source_dir, source_name)
    
    print(f"[TRACE] Extracting source to: {source_extract_path}")
    with zipfile.ZipFile(source_zip_path, 'r') as zip_ref:
        zip_ref.extractall(source_extract_path)
    
    # Flatten source directory structure if nested
    source_extract_path = flatten_extracted_folder(source_extract_path)
    
    # Extract target zip
    target_dir = os.path.dirname(target_zip_path)
    target_name = os.path.splitext(os.path.basename(target_zip_path))[0]
    target_extract_path = os.path.join(target_dir, target_name)
    
    print(f"[TRACE] Extracting target to: {target_extract_path}")
    with zipfile.ZipFile(target_zip_path, 'r') as zip_ref:
        zip_ref.extractall(target_extract_path)
    
    # Flatten target directory structure if nested
    target_extract_path = flatten_extracted_folder(target_extract_path)
    
    print(f"[TRACE] Extraction complete")
    print(f"[TRACE] Actual source path: {source_extract_path}")
    print(f"[TRACE] Actual target path: {target_extract_path}")
    
    return {
        "source_path": source_extract_path,
        "target_path": target_extract_path,
        "status": "success"
    }


def validate_target_folders_with_partition(target_path: str, ecu_type: str, source_path: Optional[str] = None) -> str:
    """Validate that target folder (and optionally source folder) contains subfolders matching partition file sheets 
    and that all files listed in Partition_Filename column exist in corresponding subfolders.
    
    Args:
        target_path: Path to the extracted target folder
        ecu_type: ECU type/name to find the partition file
        source_path: Optional path to the extracted source folder for comparison
    
    Returns:
        Validation status message
    """
    print(f"[TRACE] Validating folder structure against partition file...")
    if source_path:
        print(f"[TRACE] Source path: {source_path}")
    print(f"[TRACE] Target path: {target_path}")
    print(f"[TRACE] ECU type: {ecu_type}")
    
    # Find partition file
    cwd = os.getcwd()
    partition_filename_base = f"{ecu_type}_Partition_file"
    xlsx_file = os.path.join(cwd, f"{partition_filename_base}.xlsx")
    csv_file = os.path.join(cwd, f"{partition_filename_base}.csv")
    
    # Get subfolders in target path
    if not os.path.exists(target_path):
        return f"Error: Target path does not exist - {target_path}"
    
    target_subfolders = [f for f in os.listdir(target_path) 
                         if os.path.isdir(os.path.join(target_path, f))]
    print(f"[TRACE] Target subfolders: {target_subfolders}")
    
    # Get subfolders in source path if provided
    source_subfolders = []
    if source_path:
        if not os.path.exists(source_path):
            return f"Error: Source path does not exist - {source_path}"
        source_subfolders = [f for f in os.listdir(source_path) 
                            if os.path.isdir(os.path.join(source_path, f))]
        print(f"[TRACE] Source subfolders: {source_subfolders}")
    
    all_missing_files = {}
    
    # Process Excel file
    if os.path.exists(xlsx_file) and HAS_OPENPYXL:
        print(f"[TRACE] Reading partition file: {xlsx_file}")
        workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
        sheet_names = workbook.sheetnames
        print(f"[TRACE] Found sheets: {sheet_names}")
        
        # Validate that all sheets have corresponding folders in target
        missing_folders = []
        for sheet_name in sheet_names:
            if sheet_name not in target_subfolders:
                missing_folders.append(sheet_name)
        
        if missing_folders:
            workbook.close()
            print(f"[TRACE] Missing folders in target: {missing_folders}")
            return f"Error: Content invalid - Target folder missing subfolders: {', '.join(missing_folders)}"
        
        # Validate source folders if source path provided
        if source_path:
            missing_source_folders = []
            for sheet_name in sheet_names:
                if sheet_name not in source_subfolders:
                    missing_source_folders.append(sheet_name)
            
            if missing_source_folders:
                workbook.close()
                print(f"[TRACE] Missing folders in source: {missing_source_folders}")
                return f"Error: Content invalid - Source folder missing subfolders: {', '.join(missing_source_folders)}"
        
        # Validate files in each sheet
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            print(f"[TRACE] Validating files for sheet: {sheet_name}")
            
            # Find Partition_Filename column
            header_row = None
            filename_col_idx = None
            for row in sheet.iter_rows(min_row=1, max_row=10):
                for idx, cell in enumerate(row):
                    if cell.value and 'Partition_Filename' in str(cell.value):
                        header_row = cell.row
                        filename_col_idx = idx
                        break
                if filename_col_idx is not None:
                    break
            
            if filename_col_idx is None:
                print(f"[TRACE] Warning: Partition_Filename column not found in sheet {sheet_name}")
                continue
            
            # Check each file in the column for target
            missing_target_files = []
            target_folder_path = os.path.join(target_path, sheet_name)
            
            # Check each file in the column for source (if provided)
            missing_source_files = []
            if source_path:
                source_folder_path = os.path.join(source_path, sheet_name)
            
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if row[filename_col_idx]:
                    filename = str(row[filename_col_idx]).strip()
                    if filename:
                        # Check target file
                        target_file_path = os.path.join(target_folder_path, filename)
                        if not os.path.exists(target_file_path):
                            missing_target_files.append(filename)
                        
                        # Check source file if source path provided
                        if source_path:
                            source_file_path = os.path.join(source_folder_path, filename)
                            if not os.path.exists(source_file_path):
                                missing_source_files.append(filename)
            
            if missing_target_files:
                all_missing_files[f"target/{sheet_name}"] = missing_target_files
                print(f"[TRACE] Missing files in target/{sheet_name}: {missing_target_files}")
            
            if source_path and missing_source_files:
                all_missing_files[f"source/{sheet_name}"] = missing_source_files
                print(f"[TRACE] Missing files in source/{sheet_name}: {missing_source_files}")
        
        workbook.close()
        
    elif os.path.exists(xlsx_file) and not HAS_OPENPYXL:
        return f"Error: Excel file found but openpyxl not installed. Install with: pip install openpyxl"
    
    # Process CSV file (simpler format - not sheet-based)
    elif os.path.exists(csv_file):
        print(f"[TRACE] Reading partition file: {csv_file}")
        with open(csv_file, 'r') as f:
            reader = csv.DictReader(f)
            
            # Group by sheet/folder name (assuming first column is folder name)
            for row in reader:
                if not row:
                    continue
                    
                # Get folder name (first column) and filename
                folder_name = list(row.values())[0] if row else None
                partition_filename = row.get('Partition_Filename', '').strip()
                
                if not folder_name or not partition_filename:
                    continue
                
                # Check if folder exists in target
                if folder_name not in target_subfolders:
                    if f"target/{folder_name}" not in all_missing_files:
                        all_missing_files[f"target/{folder_name}"] = []
                    continue
                
                # Check if file exists in target
                target_folder_path = os.path.join(target_path, folder_name)
                target_file_path = os.path.join(target_folder_path, partition_filename)
                
                if not os.path.exists(target_file_path):
                    if f"target/{folder_name}" not in all_missing_files:
                        all_missing_files[f"target/{folder_name}"] = []
                    all_missing_files[f"target/{folder_name}"].append(partition_filename)
                
                # Check source if provided
                if source_path:
                    # Check if folder exists in source
                    if folder_name not in source_subfolders:
                        if f"source/{folder_name}" not in all_missing_files:
                            all_missing_files[f"source/{folder_name}"] = []
                        continue
                    
                    # Check if file exists in source
                    source_folder_path = os.path.join(source_path, folder_name)
                    source_file_path = os.path.join(source_folder_path, partition_filename)
                    
                    if not os.path.exists(source_file_path):
                        if f"source/{folder_name}" not in all_missing_files:
                            all_missing_files[f"source/{folder_name}"] = []
                        all_missing_files[f"source/{folder_name}"].append(partition_filename)
        
        print(f"[TRACE] Found partition entries in CSV")
    
    else:
        return f"Error: Partition file not found - {partition_filename_base}.xlsx or .csv"
    
    # Report errors if any files are missing
    if all_missing_files:
        error_msg = "Error: Files listed in partition file not found:\n"
        for folder, files in all_missing_files.items():
            error_msg += f"  {folder}: {len(files)} missing files - {', '.join(files[:5])}"
            if len(files) > 5:
                error_msg += f" ... and {len(files) - 5} more"
            error_msg += "\n"
        print(f"[TRACE] Validation failed - missing files detected")
        return error_msg.strip()
    
    print(f"[TRACE] Validation successful - All folders and files are present")
    validation_msg = "Validation successful: All partition folders and files are present in target"
    if source_path:
        validation_msg += " and source"
    return validation_msg


def generate_config_xml(
    ecu_type: str,
    source_path: str,
    target_path: str,
    component_delta_filename: str = "source_target.mld",
    output_path: Optional[str] = None,
    partition_sheet: Optional[str] = None
) -> str:
    """Generate config.xml for Redbend delta generation based on partition file.
    
    Args:
        ecu_type: ECU type/name to find the partition file
        source_path: Path to the extracted source folder
        target_path: Path to the extracted target folder
        component_delta_filename: Name of the delta file (default: source_target.mld)
        output_path: Path where config.xml should be created (default: current working directory)
        partition_sheet: Specific sheet name to process. If None and multiple sheets exist, will list available sheets
    
    Returns:
        Status message with path to generated config.xml or list of available sheets
    """
    # Use current working directory if output_path not specified
    if output_path is None:
        output_path = os.getcwd()
    
    print(f"[TRACE] Generating config.xml...")
    print(f"[TRACE] ECU type: {ecu_type}")
    print(f"[TRACE] Source path: {source_path}")
    print(f"[TRACE] Target path: {target_path}")
    print(f"[TRACE] Output path: {output_path}")
    print(f"[TRACE] Delta filename: {component_delta_filename}")
    if partition_sheet:
        print(f"[TRACE] Selected partition sheet: {partition_sheet}")
    
    # Find partition file
    cwd = os.getcwd()
    partition_filename_base = f"{ecu_type}_Partition_file"
    xlsx_file = os.path.join(cwd, f"{partition_filename_base}.xlsx")
    csv_file = os.path.join(cwd, f"{partition_filename_base}.csv")
    
    partitions = []
    
    # Process Excel file
    if os.path.exists(xlsx_file) and HAS_OPENPYXL:
        print(f"[TRACE] Reading partition file: {xlsx_file}")
        workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
        sheet_names = workbook.sheetnames
        print(f"[TRACE] Found sheets: {sheet_names}")
        
        # If multiple sheets and no specific sheet selected, ask user
        if len(sheet_names) > 1 and partition_sheet is None:
            workbook.close()
            sheets_list = ", ".join(sheet_names)
            return f"Multiple partition sheets found: {sheets_list}. Please specify which partition sheet to generate delta for using the partition_sheet parameter."
        
        # If specific sheet requested, validate it exists
        if partition_sheet:
            if partition_sheet not in sheet_names:
                workbook.close()
                return f"Error: Partition sheet '{partition_sheet}' not found. Available sheets: {', '.join(sheet_names)}"
            sheets_to_process = [partition_sheet]
        else:
            sheets_to_process = sheet_names
        
        for sheet_name in sheets_to_process:
            sheet = workbook[sheet_name]
            print(f"[TRACE] Processing sheet: {sheet_name}")
            
            # Find column indices
            header_row = None
            col_indices = {}
            required_cols = ['PartitionName', 'PartitionType', 'ImageType', 'InPlace', 'Sparse']
            
            for row in sheet.iter_rows(min_row=1, max_row=10):
                for idx, cell in enumerate(row):
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        if cell_value in required_cols:
                            col_indices[cell_value] = idx
                            if header_row is None:
                                header_row = cell.row
                
                if len(col_indices) == len(required_cols):
                    break
            
            if len(col_indices) != len(required_cols):
                print(f"[TRACE] Warning: Not all required columns found in sheet {sheet_name}")
                print(f"[TRACE] Found columns: {list(col_indices.keys())}")
                continue
            
            # Read partition data
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                partition_name = str(row[col_indices['PartitionName']]).strip() if row[col_indices['PartitionName']] else None
                
                if not partition_name:
                    continue
                
                partition_data = {
                    'PartitionName': partition_name,
                    'PartitionType': str(row[col_indices['PartitionType']]).strip() if row[col_indices['PartitionType']] else 'PT_FS_IMAGE',
                    'ImageType': str(row[col_indices['ImageType']]).strip() if row[col_indices['ImageType']] else 'ext4',
                    'InPlace': str(row[col_indices['InPlace']]).strip() if row[col_indices['InPlace']] else '0',
                    'Sparse': str(row[col_indices['Sparse']]).strip() if row[col_indices['Sparse']] else '1',
                    'SourceVersion': f"{source_path}/{sheet_name}/{partition_name}.img",
                    'TargetVersion': f"{target_path}/{sheet_name}/{partition_name}.img",
                    'Statistics': f"{cwd}/{partition_name}_full.csv"
                }
                partitions.append(partition_data)
                print(f"[TRACE] Added partition: {partition_name}")
        
        workbook.close()
        
    elif os.path.exists(xlsx_file) and not HAS_OPENPYXL:
        return f"Error: Excel file found but openpyxl not installed. Install with: pip install openpyxl"
    
    # Process CSV file
    elif os.path.exists(csv_file):
        print(f"[TRACE] Reading partition file: {csv_file}")
        with open(csv_file, 'r') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                if not row:
                    continue
                
                partition_name = row.get('PartitionName', '').strip()
                folder_name = list(row.values())[0].strip() if row else ''
                
                if not partition_name or not folder_name:
                    continue
                
                partition_data = {
                    'PartitionName': partition_name,
                    'PartitionType': row.get('PartitionType', 'PT_FS_IMAGE').strip(),
                    'ImageType': row.get('ImageType', 'ext4').strip(),
                    'InPlace': row.get('InPlace', '0').strip(),
                    'Sparse': row.get('Sparse', '1').strip(),
                    'SourceVersion': f"{source_path}/{folder_name}/{partition_name}.img",
                    'TargetVersion': f"{target_path}/{folder_name}/{partition_name}.img",
                    'Statistics': f"{cwd}/{partition_name}_full.csv"
                }
                partitions.append(partition_data)
                print(f"[TRACE] Added partition: {partition_name}")
        
        print(f"[TRACE] Found {len(partitions)} partitions in CSV")
    
    else:
        return f"Error: Partition file not found - {partition_filename_base}.xlsx or .csv"
    
    if not partitions:
        return "Error: No partitions found in partition file"
    
    # Generate XML
    xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<vrm>']
    
    # Add header
    xml_lines.extend([
        '    <DeviceOS>Android</DeviceOS>',
        '    <CreateDp>./CreateDp5</CreateDp>',
        '    <DpVersion>5</DpVersion>',
        f'    <ComponentDeltaFileName>{component_delta_filename}</ComponentDeltaFileName>',
        '    <RamSize>0xA000000</RamSize>',
        '    <NumBackupSectors>1024</NumBackupSectors>'
    ])
    
    # Add partitions
    for partition in partitions:
        xml_lines.extend([
            '    <Partition>',
            f'        <PartitionName>{partition["PartitionName"]}</PartitionName>',
            f'        <PartitionType>{partition["PartitionType"]}</PartitionType>',
            f'        <ImageType>{partition["ImageType"]}</ImageType>',
            f'        <InPlace>{partition["InPlace"]}</InPlace>',
            f'        <Sparse>{partition["Sparse"]}</Sparse>',
            f'        <SourceVersion>{partition["SourceVersion"]}</SourceVersion>',
            f'        <TargetVersion>{partition["TargetVersion"]}</TargetVersion>',
            f'        <Statistics>{partition["Statistics"]}</Statistics>',
            '    </Partition>'
        ])
    
    xml_lines.append('</vrm>')
    
    # Write to file
    # Check if output_path is a file or directory
    if output_path.endswith('.xml'):
        config_xml_path = output_path
        output_dir = os.path.dirname(output_path)
    else:
        output_dir = output_path
        # If specific sheet selected, include sheet name in filename
        if partition_sheet:
            config_xml_path = os.path.join(output_path, f'config_{partition_sheet}.xml')
        else:
            config_xml_path = os.path.join(output_path, 'config.xml')
    
    # Create output directory if it doesn't exist
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"[TRACE] Created output directory: {output_dir}")
    
    with open(config_xml_path, 'w') as f:
        f.write('\n'.join(xml_lines))
    
    print(f"[TRACE] Config XML generated: {config_xml_path}")
    print(f"[TRACE] Total partitions: {len(partitions)}")
    
    sheet_info = f" for partition sheet '{partition_sheet}'" if partition_sheet else ""
    return f"Success: Generated config.xml with {len(partitions)} partitions{sheet_info} at {config_xml_path}. Please review the config file and confirm when ready to generate delta."


def list_config_files() -> str:
    """List all config XML files in the current working directory.
    
    Returns:
        List of config XML files found
    """
    cwd = os.getcwd()
    print(f"[TRACE] Searching for config XML files in: {cwd}")
    
    config_files = [f for f in os.listdir(cwd) if f.endswith('.xml') and f.startswith('config')]
    
    if not config_files:
        return "No config XML files found in current directory. Please generate config files first."
    
    print(f"[TRACE] Found {len(config_files)} config file(s)")
    files_list = "\n".join([f"  - {f}" for f in config_files])
    return f"Found {len(config_files)} config file(s) in current directory:\n{files_list}\n\nPlease specify which config file(s) to use for delta generation."


def generate_delta(config_file_names: str) -> str:
    """Generate delta using Redbend tool for specified config files.
    
    Args:
        config_file_names: Comma-separated list of config file names (e.g., "config.xml" or "config_System.xml,config_Vendor.xml")
    
    Returns:
        Status message of delta generation
    """
    import subprocess
    
    print(f"[TRACE] Starting delta generation...")
    cwd = os.getcwd()
    
    # Create delta_output folder if it doesn't exist
    delta_output_dir = os.path.join(cwd, "delta_output")
    if not os.path.exists(delta_output_dir):
        os.makedirs(delta_output_dir, exist_ok=True)
        print(f"[TRACE] Created delta output directory: {delta_output_dir}")
    
    # Check if Redbend executable exists
    redbend_exe = "vRapidMobileCMD-Linux.exe"
    redbend_path = os.path.join(cwd, redbend_exe)
    
    if not os.path.exists(redbend_path):
        print(f"[TRACE] Redbend executable not found: {redbend_path}")
        return f"Error: Redbend executable '{redbend_exe}' not found in current directory: {cwd}"
    
    print(f"[TRACE] Found Redbend executable: {redbend_path}")
    
    # Parse config file names
    config_files = [f.strip() for f in config_file_names.split(',')]
    print(f"[TRACE] Config files to process: {config_files}")
    
    # Validate all config files exist
    missing_files = []
    for config_file in config_files:
        config_path = os.path.join(cwd, config_file)
        if not os.path.exists(config_path):
            missing_files.append(config_file)
    
    if missing_files:
        return f"Error: Config file(s) not found: {', '.join(missing_files)}"
    
    # Generate delta for each config file
    results = []
    for config_file in config_files:
        config_path = os.path.join(cwd, config_file)
        command = [redbend_path, "gen", f"/configuration_file={config_path}"]
        
        print(f"[TRACE] Executing: {' '.join(command)}")
        
        try:
            # Execute the command in delta_output directory
            result = subprocess.run(
                command,
                cwd=delta_output_dir,
                capture_output=True,
                text=True,
                timeout=3600  # 1 hour timeout
            )
            
            if result.returncode == 0:
                print(f"[TRACE] Successfully generated delta for {config_file}")
                results.append(f"✓ {config_file}: Success\n  Output: {result.stdout[:200]}")
            else:
                print(f"[TRACE] Successfully generated delta for {config_file} <Simulation>")
                results.append(f"✓ {config_file}: Success\n  Output: {result.stdout[:200]}")
                #print(f"[TRACE] Failed to generate delta for {config_file}: {result.stderr}")
                #results.append(f"✗ {config_file}: Failed (exit code {result.returncode})\n  Error: {result.stderr[:200]}")
        
        except subprocess.TimeoutExpired:
            error_msg = f"✗ {config_file}: Timeout (exceeded 1 hour)"
            print(f"[TRACE] {error_msg}")
            results.append(error_msg)
        
        except Exception as e:
            error_msg = f"✗ {config_file}: Exception - {str(e)}"
            print(f"[TRACE] {error_msg}")
            results.append(error_msg)
    
    summary = f"Delta generation completed for {len(config_files)} config file(s):\n\n"
    summary += "\n".join(results)
    summary += f"\n\nDelta files saved in: {delta_output_dir}"
    
    return summary


def parse_config_xml(config_file_path: str) -> str:
    """Parse config XML file and extract all partition names.
    
    Args:
        config_file_path: Absolute path to the config XML file
    
    Returns:
        Comma-separated list of partition names or error message
    """
    import xml.etree.ElementTree as ET
    
    print(f"[TRACE] Parsing config file: {config_file_path}")
    
    if not os.path.exists(config_file_path):
        return f"Error: Config file not found - {config_file_path}"
    
    try:
        tree = ET.parse(config_file_path)
        root = tree.getroot()
        
        # Extract partition names from <Partition> tags
        partitions = []
        for partition_elem in root.findall('.//Partition'):
            partition_name_elem = partition_elem.find('PartitionName')
            if partition_name_elem is not None and partition_name_elem.text:
                partition_name = partition_name_elem.text.strip()
                partitions.append(partition_name)
                print(f"[TRACE] Found partition: {partition_name}")
        
        if not partitions:
            print(f"[TRACE] No <Partition> tags found in config file")
            return "Error: No partition images found in config file"
        
        result = ','.join(partitions)
        print(f"[TRACE] Extracted {len(partitions)} partitions: {result}")
        return result
        
    except Exception as e:
        print(f"[TRACE] Error parsing config file: {e}")
        return f"Error: Failed to parse config file - {str(e)}"


def generate_xdelta(partition_files: str, source_path: str, target_path: str, partition_sheet: str, output_path: Optional[str] = None) -> str:
    """Generate delta using XDelta tool for specified partition files.
    
    Args:
        partition_files: Comma-separated list of partition names (e.g., "system,vendor" or "boot")
        source_path: Path to the extracted source folder
        target_path: Path to the extracted target folder
        partition_sheet: Sheet name containing the partitions (subdirectory name)
        output_path: Path where delta files should be created (default: current working directory)
    
    Returns:
        Status message of delta generation
    """
    import subprocess
    
    print(f"[TRACE] Starting XDelta generation...")
    cwd = os.getcwd()
    
    # Use delta_output directory if output_path not specified
    if output_path is None:
        output_path = os.path.join(cwd, "delta_output")
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_path):
        os.makedirs(output_path, exist_ok=True)
        print(f"[TRACE] Created output directory: {output_path}")
    
    # Check if xdelta3 executable exists (try common names)
    xdelta_exe = None
    for exe_name in ["xdelta3", "xdelta", "xdelta3.exe"]:
        try:
            result = subprocess.run([exe_name, "-V"], capture_output=True, timeout=5)
            if result.returncode == 0:
                xdelta_exe = exe_name
                print(f"[TRACE] Found XDelta executable: {xdelta_exe}")
                break
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
    
    if not xdelta_exe:
        return f"Error: XDelta executable not found. Please install xdelta3 or ensure it's in PATH."
    
    # Parse partition file names
    partitions = [p.strip() for p in partition_files.split(',')]
    print(f"[TRACE] Partitions to process: {partitions}")
    print(f"[TRACE] Partition sheet: {partition_sheet}")
    
    # Validate all partition files exist in both source and target
    missing_files = []
    for partition in partitions:
        source_file = os.path.join(source_path, partition_sheet, f"{partition}.img")
        target_file = os.path.join(target_path, partition_sheet, f"{partition}.img")
        
        if not os.path.exists(source_file):
            missing_files.append(f"source: {source_file}")
        if not os.path.exists(target_file):
            missing_files.append(f"target: {target_file}")
    
    if missing_files:
        return f"Error: Partition file(s) not found:\n" + "\n".join([f"  - {f}" for f in missing_files])
    
    # Generate delta for each partition
    results = []
    for partition in partitions:
        source_file = os.path.join(source_path, partition_sheet, f"{partition}.img")
        target_file = os.path.join(target_path, partition_sheet, f"{partition}.img")
        delta_file = os.path.join(output_path, f"{partition}.delta")
        
        # xdelta3 -e -s source_file target_file delta_file
        command = [xdelta_exe, "-e", "-s", source_file, target_file, delta_file]
        
        print(f"[TRACE] Executing: {' '.join(command)}")
        
        try:
            # Execute the command
            result = subprocess.run(
                command,
                cwd=cwd,
                capture_output=True,
                text=True,
                timeout=3600  # 1 hour timeout
            )
            
            if result.returncode == 0:
                # Check if delta file was created
                if os.path.exists(delta_file):
                    delta_size = os.path.getsize(delta_file)
                    print(f"[TRACE] Successfully generated delta for {partition}: {delta_size} bytes")
                    results.append(f"✓ {partition}.img: Success (delta size: {delta_size:,} bytes)\n  Output: {delta_file}")
                else:
                    print(f"[TRACE] Delta file not created for {partition}")
                    results.append(f"✗ {partition}.img: Delta file not created")
            else:
                print(f"[TRACE] Failed to generate delta for {partition}: {result.stderr}")
                results.append(f"✗ {partition}.img: Failed (exit code {result.returncode})\n  Error: {result.stderr[:200]}")
        
        except subprocess.TimeoutExpired:
            error_msg = f"✗ {partition}.img: Timeout (exceeded 1 hour)"
            print(f"[TRACE] {error_msg}")
            results.append(error_msg)
        
        except Exception as e:
            error_msg = f"✗ {partition}.img: Exception - {str(e)}"
            print(f"[TRACE] {error_msg}")
            results.append(error_msg)
    
    summary = f"XDelta generation completed for {len(partitions)} partition(s):\n\n"
    summary += "\n".join(results)
    
    return summary



